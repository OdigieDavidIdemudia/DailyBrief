import io
import csv
from datetime import datetime
from fpdf import FPDF
from app.db import DailyLogModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter

def safe_text(text: str) -> str:
    if not text:
        return ""
    return str(text).encode('latin-1', 'replace').decode('latin-1')

# 1. CSV Report Generator
def generate_csv_report(logs: list, date_str: str) -> str:
    # Calculate summary metrics
    total = len(logs)
    completed = len([l for l in logs if l.status == 'Completed'])
    in_progress = len([l for l in logs if l.status in ['In Progress', 'Pending']])
    blocked = len([l for l in logs if l.status in ['Blocked', 'Flagged']])

    output = io.StringIO()
    
    # Write # Header block
    output.write("# Header block\n")
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Project Name', 'Daily Brief'])
    writer.writerow(['Report Date', date_str])
    writer.writerow(['Owner', 'David Odigie'])
    output.write("\n")
    
    # Write # Summary block
    output.write("# Summary block\n")
    writer.writerow(['Total Tasks', total])
    writer.writerow(['Completed', completed])
    writer.writerow(['In Progress', in_progress])
    writer.writerow(['Blocked', blocked])
    output.write("\n")
    
    # Write # Tasks block
    output.write("# Tasks block\n")
    writer.writerow([
        'Date',
        'Task_ID',
        'Category',
        'Task_Title',
        'Priority',
        'Status',
        'Summary',
        'Challenges',
        'Mail_Trail',
        'Critical_Blocker'
    ])
    
    # Rows
    for idx, log in enumerate(logs):
        task_id = str(idx + 1)
        # Resolve joins
        title = log.blueprint.title if log.blueprint else "Deleted Task"
        category = log.blueprint.category if log.blueprint else "Daily"
        priority = log.blueprint.priority if log.blueprint else "Standard"
        
        writer.writerow([
            log.date,
            task_id,
            category,
            title,
            priority,
            log.status,
            log.summary or '',
            log.challenges or '',
            log.mail_trail or '',
            'TRUE' if log.is_critical else 'FALSE'
        ])
        
    return output.getvalue()

# 2. PDF Report Generator (using fpdf2)
class DailyReportPDF(FPDF):
    def __init__(self, date_str: str):
        super().__init__(orientation='P', unit='mm', format='A4')
        self.date_str = date_str
        self.set_margins(15, 15, 15)
        self.set_auto_page_break(auto=True, margin=20)
        
    def header(self):
        pass

    def footer(self):
        self.set_y(-15)
        # Draw outline divider line
        self.set_draw_color(226, 232, 240) # Border gray
        self.set_line_width(0.2)
        self.line(15, self.get_y(), 195, self.get_y())
        
        self.set_y(-12)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(100, 116, 139) # Slate 500
        
        # Left footer note
        self.cell(0, 10, 'End of Daily Report', align='L')
        # Right page number
        self.set_x(-35)
        self.cell(20, 10, f'Page {self.page_no()}', align='R')

def generate_pdf_report(logs: list, date_str: str) -> bytes:
    pdf = DailyReportPDF(date_str)
    pdf.add_page()
    
    # Colors
    c_primary = (15, 23, 42)      # Slate 900
    c_secondary = (100, 116, 139)  # Slate 500
    c_accent = (30, 53, 208)      # Accent primary blue
    c_background = (248, 249, 255) # Soft background
    c_border = (226, 232, 240)    # Border gray
    c_red = (186, 26, 26)         # Critical red
    c_green = (27, 94, 32)        # Success green
    
    # Format human-friendly date: e.g. "June 21, 2026"
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        human_date = dt.strftime("%B %d, %Y")
    except Exception:
        human_date = date_str

    # 1. Header Title
    pdf.set_y(18)
    pdf.set_font('Helvetica', 'B', 15)
    pdf.set_text_color(*c_accent)
    pdf.cell(0, 10, 'Daily Brief', align='C', new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(*c_secondary)
    pdf.cell(0, 5, f"Report Date: {date_str} | Owner: David Odigie", align='C', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    
    # 2. Executive Metrics Summary Box
    total = len(logs)
    completed = len([l for l in logs if l.status == 'Completed'])
    in_progress = len([l for l in logs if l.status in ['In Progress', 'Pending']])
    critical = len([l for l in logs if l.status in ['Blocked', 'Flagged']])
    
    # Metric values and labels
    metrics = [
        {"val": str(total), "label": "Total Tasks"},
        {"val": str(completed), "label": "Completed"},
        {"val": str(in_progress), "label": "In Progress"},
        {"val": str(critical), "label": "Blocked / Critical"}
    ]
    
    # Draw Background box for metrics (width: 180mm, height: 18mm)
    start_y = pdf.get_y()
    pdf.set_fill_color(*c_background)
    pdf.set_draw_color(*c_border)
    pdf.rect(15, start_y, 180, 18, style='FD')
    
    col_width = 180 / 4
    for idx, m in enumerate(metrics):
        x_pos = 15 + (idx * col_width)
        
        # Draw metric value
        pdf.set_xy(x_pos, start_y + 2)
        pdf.set_font('Helvetica', 'B', 14)
        if idx == 3 and critical > 0:
            pdf.set_text_color(*c_red)
        else:
            pdf.set_text_color(*c_primary)
        pdf.cell(col_width, 6, m["val"], align='C')
        
        # Draw metric label
        pdf.set_xy(x_pos, start_y + 9)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*c_secondary)
        pdf.cell(col_width, 5, m["label"], align='C')
        
    pdf.set_y(start_y + 23)
    
    # 3. Critical Blockers Section (only if there are critical tasks)
    critical_logs = [l for l in logs if l.status in ['Blocked', 'Flagged'] or l.is_critical]
    if critical_logs:
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(*c_red)
        pdf.cell(0, 6, 'CRITICAL BLOCKERS', new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)
        
        for log in critical_logs:
            title = safe_text(log.blueprint.title) if log.blueprint else "Deleted Task"
            priority = safe_text(log.blueprint.priority) if log.blueprint else "Critical"
            challenges = safe_text(log.challenges) if log.challenges else "No explanation provided."
            
            # Estimate height needed
            text_lines = len(pdf.multi_cell(170, 4, f"Challenges: {challenges}", dry_run=True, output="LINES"))
            card_height = 8 + (text_lines * 4.5)
            
            # Avoid orphans/widows
            if pdf.get_y() + card_height > 270:
                pdf.add_page()
                
            card_start_y = pdf.get_y()
            # Draw Card Background (soft red tint)
            pdf.set_fill_color(255, 248, 248)
            pdf.set_draw_color(*c_red)
            pdf.set_line_width(0.15)
            pdf.rect(15, card_start_y, 180, card_height, style='FD')
            
            # Title of critical task
            pdf.set_xy(20, card_start_y + 2)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_text_color(*c_primary)
            pdf.cell(100, 5, f"- [{priority}] {title}")
            
            # Status tag right aligned
            pdf.set_xy(145, card_start_y + 2)
            pdf.set_text_color(*c_red)
            pdf.cell(45, 5, f"STATUS: {log.status.upper()}", align='R')
            
            # Challenges
            pdf.set_xy(20, card_start_y + 7)
            pdf.set_font('Helvetica', '', 8.5)
            pdf.set_text_color(*c_secondary)
            pdf.multi_cell(170, 4.5, f"Challenges: {challenges}")
            
            pdf.set_y(card_start_y + card_height + 3)
            
        pdf.ln(4)

    # 4. Detailed Task Breakdown
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_text_color(*c_primary)
    pdf.cell(0, 6, 'TASK BREAKDOWN BY STATUS', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    
    status_groups = [
        {"name": "Completed Operations", "statuses": ["Completed"], "color": c_green},
        {"name": "In Progress", "statuses": ["In Progress"], "color": c_accent},
        {"name": "Pending / Blocked / Other", "statuses": ["Pending", "Flagged", "Blocked"], "color": c_secondary}
    ]
    
    for group in status_groups:
        group_logs = [l for l in logs if l.status in group["statuses"]]
        if not group_logs:
            continue
            
        # Draw Group Header
        if pdf.get_y() > 260:
            pdf.add_page()
            
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(*group["color"])
        pdf.cell(0, 5, group["name"], new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)
        
        for log in group_logs:
            title = safe_text(log.blueprint.title) if log.blueprint else "Deleted Task"
            priority = safe_text(log.blueprint.priority) if log.blueprint else "Standard"
            summary = safe_text(log.summary) if log.summary else "No summary provided."
            challenges = safe_text(log.challenges)
            mail_trail = safe_text(log.mail_trail)
            
            # Estimate card height
            sum_lines = len(pdf.multi_cell(170, 4, f"Summary: {summary}", dry_run=True, output="LINES"))
            chal_lines = 0
            if challenges:
                chal_lines = len(pdf.multi_cell(170, 4, f"Challenges: {challenges}", dry_run=True, output="LINES"))
            mail_lines = 0
            if mail_trail:
                mail_lines = len(pdf.multi_cell(170, 4, f"Mail Trail: {mail_trail}", dry_run=True, output="LINES"))
                
            total_text_lines = sum_lines + chal_lines + mail_lines
            card_height = 8 + (total_text_lines * 4) + (2 if challenges else 0) + (2 if mail_trail else 0)
            
            if pdf.get_y() + card_height > 270:
                pdf.add_page()
                
            card_start_y = pdf.get_y()
            # Draw light grey card border
            pdf.set_fill_color(255, 255, 255)
            pdf.set_draw_color(*c_border)
            pdf.set_line_width(0.1)
            pdf.rect(15, card_start_y, 180, card_height, style='FD')
            
            # Title & Priority
            pdf.set_xy(18, card_start_y + 2)
            pdf.set_font('Helvetica', 'B', 8.5)
            pdf.set_text_color(*c_primary)
            pdf.cell(100, 4, f"[{priority}] {title}")
            
            # Status right aligned
            pdf.set_xy(145, card_start_y + 2)
            pdf.set_text_color(*group["color"])
            pdf.cell(45, 4, log.status, align='R')
            
            # Summary & Challenges text
            pdf.set_xy(18, card_start_y + 6)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*c_primary)
            pdf.multi_cell(174, 4, f"Summary: {summary}")
            
            if challenges:
                pdf.set_x(18)
                pdf.set_text_color(*c_secondary)
                pdf.multi_cell(174, 4, f"Challenges: {challenges}")
                
            if mail_trail:
                pdf.set_x(18)
                pdf.set_text_color(80, 80, 100) # Slightly different text color for mail
                pdf.multi_cell(174, 4, f"Mail Trail: {mail_trail}")
                
            pdf.set_y(card_start_y + card_height + 2)
            
        pdf.ln(3)
        
    return pdf.output()

# 3. Excel Report Generator (using openpyxl)
def generate_xlsx_report(logs: list, date_str: str) -> bytes:
    wb = Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Daily Brief
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Daily Brief"
    ws1.views.sheetView[0].showGridLines = True
    
    # Font definitions
    font_family = "Segoe UI"
    f_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    f_subtitle = Font(name=font_family, size=10, italic=True, color="E2E8F0")
    f_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_bold = Font(name=font_family, size=10, bold=True)
    f_regular = Font(name=font_family, size=10)
    f_kpi_val = Font(name=font_family, size=18, bold=True, color="1E3A8A")
    f_kpi_lbl = Font(name=font_family, size=9, bold=True, color="6B7280")
    
    # Fills
    fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_kpi = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    # Status/Priority Colors
    fill_completed = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # light green
    font_completed = Font(name=font_family, size=9, bold=True, color="065F46")
    
    fill_progress = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # light blue
    font_progress = Font(name=font_family, size=9, bold=True, color="1E40AF")
    
    fill_pending = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # light gray
    font_pending = Font(name=font_family, size=9, bold=True, color="374151")
    
    fill_blocked = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
    font_blocked = Font(name=font_family, size=9, bold=True, color="991B1B")

    # Priority Colors
    fill_high = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # light orange
    font_high = Font(name=font_family, size=9, bold=True, color="9A3412")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # 1. Header Block (Navy banner)
    ws1.merge_cells("A1:I1")
    ws1.merge_cells("A2:I2")
    ws1["A1"] = "Daily Brief"
    ws1["A1"].font = f_title
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws1["A2"] = f"Report Date: {date_str}  |  Owner: David Odigie"
    ws1["A2"].font = f_subtitle
    ws1["A2"].fill = fill_navy
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 20

    # 2. KPI Summary Cards
    total = len(logs)
    completed = len([l for l in logs if l.status == 'Completed'])
    in_progress = len([l for l in logs if l.status in ['In Progress', 'Pending']])
    blocked = len([l for l in logs if l.status in ['Blocked', 'Flagged']])
    
    kpi_cols = [("B", "Total Tasks", total), ("D", "Completed", completed), ("F", "In Progress", in_progress), ("H", "Blocked", blocked)]
    
    ws1.row_dimensions[4].height = 25
    ws1.row_dimensions[5].height = 15
    
    for col, label, val in kpi_cols:
        c1 = col
        c2 = chr(ord(col) + 1)
        
        ws1.merge_cells(f"{c1}4:{c2}4")
        ws1.merge_cells(f"{c1}5:{c2}5")
        
        ws1[f"{c1}4"] = val
        ws1[f"{c1}4"].font = f_kpi_val
        ws1[f"{c1}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws1[f"{c1}4"].fill = fill_kpi
        
        ws1[f"{c1}5"] = label
        ws1[f"{c1}5"].font = f_kpi_lbl
        ws1[f"{c1}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws1[f"{c1}5"].fill = fill_kpi
        
        for r in [4, 5]:
            ws1[f"{c1}{r}"].border = thin_border
            ws1[f"{c2}{r}"].border = thin_border

    # 3. Tasks Table Headers
    headers = ["Date", "Task ID", "Category", "Task Title", "Priority", "Status", "Summary", "Challenges", "Mail Trail", "Critical Blocker"]
    start_row = 7
    ws1.row_dimensions[start_row].height = 24
    
    for idx, h in enumerate(headers):
        cell = ws1.cell(row=start_row, column=idx + 1)
        cell.value = h
        cell.font = f_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center" if idx < 6 or idx == 8 else "left", vertical="center")
        cell.border = thin_border

    # 4. Tasks Table Rows
    curr_row = start_row + 1
    for idx, log in enumerate(logs):
        ws1.row_dimensions[curr_row].height = 20
        task_id = str(idx + 1)
        title = log.blueprint.title if log.blueprint else "Orphaned Task"
        category = log.blueprint.category if log.blueprint else "Daily"
        priority = log.blueprint.priority if log.blueprint else "Standard"
        
        row_fill = fill_zebra if idx % 2 == 1 else fill_white
        
        vals = [
            log.date,
            task_id,
            category,
            title,
            priority,
            log.status,
            log.summary or '',
            log.challenges or '',
            log.mail_trail or '',
            'TRUE' if log.is_critical else 'FALSE'
        ]
        
        for col_idx, val in enumerate(vals):
            cell = ws1.cell(row=curr_row, column=col_idx + 1)
            cell.value = val
            cell.font = f_regular
            cell.fill = row_fill
            cell.border = thin_border
            
            if col_idx in [0, 1, 2, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 5:
                if val == 'Completed':
                    cell.fill = fill_completed
                    cell.font = font_completed
                elif val == 'In Progress':
                    cell.fill = fill_progress
                    cell.font = font_progress
                elif val in ['Blocked', 'Flagged']:
                    cell.fill = fill_blocked
                    cell.font = font_blocked
                else:
                    cell.fill = fill_pending
                    cell.font = font_pending
            
            if col_idx == 4:
                if val in ['High', 'Critical']:
                    cell.fill = fill_high
                    cell.font = font_high
                    
        curr_row += 1
        
    # 5. Legend
    curr_row += 1
    ws1.cell(row=curr_row, column=1, value="Legend:").font = f_bold
    curr_row += 1
    
    ws1.cell(row=curr_row, column=1, value="Completed").font = font_completed
    ws1.cell(row=curr_row, column=1).fill = fill_completed
    ws1.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
    
    ws1.cell(row=curr_row, column=2, value="In Progress").font = font_progress
    ws1.cell(row=curr_row, column=2).fill = fill_progress
    ws1.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
    
    ws1.cell(row=curr_row, column=3, value="Blocked / Flagged").font = font_blocked
    ws1.cell(row=curr_row, column=3).fill = fill_blocked
    ws1.cell(row=curr_row, column=3).alignment = Alignment(horizontal="center")
    
    ws1.cell(row=curr_row, column=4, value="Pending").font = font_pending
    ws1.cell(row=curr_row, column=4).fill = fill_pending
    ws1.cell(row=curr_row, column=4).alignment = Alignment(horizontal="center")

    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ["A1", "A2", "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws1.column_dimensions["G"].width = 25
    ws1.column_dimensions["H"].width = 25

    # ----------------------------------------------------
    # Sheet 2: Raw Data
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Raw Data")
    ws2.views.sheetView[0].showGridLines = True
    
    for idx, h in enumerate(headers):
        cell = ws2.cell(row=1, column=idx + 1)
        cell.value = h
        cell.font = Font(name=font_family, size=10, bold=True)
        cell.border = thin_border
        
    for idx, log in enumerate(logs):
        task_id = str(idx + 1)
        title = log.blueprint.title if log.blueprint else "Orphaned Task"
        category = log.blueprint.category if log.blueprint else "Daily"
        priority = log.blueprint.priority if log.blueprint else "Standard"
        
        row_vals = [
            log.date,
            task_id,
            category,
            title,
            priority,
            log.status,
            log.summary or '',
            log.challenges or '',
            log.mail_trail or '',
            'TRUE' if log.is_critical else 'FALSE'
        ]
        
        for col_idx, val in enumerate(row_vals):
            cell = ws2.cell(row=idx + 2, column=col_idx + 1)
            cell.value = val
            cell.font = f_regular
            cell.border = thin_border
            
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()
